from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from rest_framework import status as drf_status
from django.db import transaction
from django.db.models import Q
from .models import SSOrder, SSOrderItem,CRMVerifiedOrderItem, CRMVerifiedOrder, Product, DispatchOrder
from django.contrib.auth import get_user_model
from .serializers import SSOrderSerializer,SS_to_CRM_Orders, CRMVerifiedOrderSerializer, CRMVerifiedOrderListSerializer, CombinedOrderTrackSerializer, SSOrderSerializerTrack, DispatchOrderSerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListAPIView
from django.shortcuts import get_object_or_404
from .utils import send_whatsapp_template
from decimal import Decimal, InvalidOperation
from django.db import transaction
from orders.models import PendingOrderItemSnapshot
from products.utils import recalculate_virtual_stock
from django.conf import settings
from products.utils import write_to_sheet
import openpyxl
from django.utils import timezone
from rest_framework.parsers import MultiPartParser
from rest_framework import status
from .models import DispatchOrder
from django.http import HttpResponse
from rest_framework.views import APIView
from datetime import datetime
import logging



logger = logging.getLogger(__name__)
User = get_user_model()


class SSOrderCreateView(APIView):
    def post(self, request):
        data = request.data

        try:
            ss_user = User.objects.get(id=data['user_id'])
            crm_user = User.objects.get(id=data['crm_id'])
            total = data['total']
            items = data['items']
            scheme_items = data.get('eligibleSchemes', [])

            # ✅ Items को Tempered और Non-Tempered में बाँटें
            tempered_items = []
            non_tempered_items = []

            for item in items:
                product = Product.objects.get(product_id=item['id'])
                sub_category = getattr(product, "sub_category", "") or ""
                if "tempered" in sub_category.lower():
                    tempered_items.append(item)
                else:
                    non_tempered_items.append(item)

            # ✅ Helper function: order create + items insert
            def create_order(order_items, label="Normal"):
                if not order_items:
                    return None

                total_amt = sum(
                    (i['price'] or 0) * i['quantity'] for i in order_items
                )

                order = SSOrder.objects.create(
                    ss_user=ss_user,
                    assigned_crm=crm_user,
                    total_amount=total_amt,
                    note=f"{label} Order"  # optional tag for clarity
                )

                for item in order_items:
                    product = Product.objects.get(product_id=item['id'])
                    SSOrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=item['quantity'],
                        price=item['price'] or 0,
                        is_scheme_item=False,
                        ss_virtual_stock=item.get('ss_virtual_stock', getattr(product, 'stock_quantity', 0))
                    )

                return order

            # ✅ Create two orders
            tempered_order = create_order(tempered_items, label="Tempered")
            normal_order = create_order(non_tempered_items, label="Accessories")

            # ✅ Scheme items — सिर्फ Non-Tempered order में add करो
            if normal_order and scheme_items:
                for reward in scheme_items:
                    product_id = (
                        reward.get('product_id') or
                        (reward.get('product', {}).get('id') if isinstance(reward.get('product'), dict) else reward.get('product'))
                    )
                    if not product_id:
                        continue

                    try:
                        product = Product.objects.get(product_id=product_id)
                        SSOrderItem.objects.create(
                            order=normal_order,
                            product=product,
                            quantity=reward.get('quantity', 0),
                            price=0,
                            is_scheme_item=True,
                            ss_virtual_stock=getattr(product, 'virtual_stock', getattr(product, 'stock_quantity', 0))
                        )
                    except Product.DoesNotExist:
                        continue

            # ✅ WhatsApp send (अब दोनों orders के लिए)
            crm_numbers = {
                2: "7678491163",
                4: "9312093178",
                7: "8595957195",
                8: "9266877089",
                9: "9266767418",
                133: "7428828836",
            }
            crm_number = crm_numbers.get(crm_user.id)
            if crm_number:
                template_name = "order_updation"
                template_language = "EN"

                for each_order in [tempered_order, normal_order]:
                    if each_order:
                        parameters = [
                            ss_user.party_name or ss_user.name,
                            str(each_order.order_id),
                            str(each_order.total_amount)
                        ]
                        send_whatsapp_template(crm_number, template_name, template_language, parameters)


            # ✅ Response
            return Response({
                "message": "Orders placed successfully.",
                "orders": {
                    "tempered_order": SSOrderSerializer(tempered_order).data if tempered_order else None,
                    "normal_order": SSOrderSerializer(normal_order).data if normal_order else None,
                }
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            print("❌ Exception occurred during order placement:")
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class SimpleSSOrderCreateView(APIView):
    """
    Only creates empty order
    No products, no items
    """

    def post(self, request):
        try:
            ss_id = request.data.get("ss_id")
            crm_id = request.data.get("crm_id")
            note = request.data.get("note", "")

            if not ss_id or not crm_id:
                return Response(
                    {"error": "ss_id and crm_id are required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            ss_user = User.objects.get(id=ss_id)
            crm_user = User.objects.get(id=crm_id)

            order = SSOrder.objects.create(
                ss_user=ss_user,
                assigned_crm=crm_user,
                total_amount=0,
                status="PENDING",
                note=note or "Empty Order"
            )

            return Response({
                "message": "Order created successfully",
                "order": SSOrderSerializer(order).data
            }, status=status.HTTP_201_CREATED)

        except User.DoesNotExist:
            return Response(
                {"error": "Invalid SS or CRM"},
                status=status.HTTP_404_NOT_FOUND
            )

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


@api_view(['POST'])
def hold_order(request, order_id):
    try:
        crm_user = request.user
        order = get_object_or_404(SSOrder, id=order_id, assigned_crm=crm_user)

        # वो snapshots लो जो पहले order verify/forward होने पर बने थे
        pending_snapshots = PendingOrderItemSnapshot.objects.filter(order=order)
        affected_products = [snap.product for snap in pending_snapshots]

        with transaction.atomic():

            # ✅ पहले snapshots delete — ये stock restore का trigger है
            pending_snapshots.delete()

            # ✅ हर product का virtual stock दोबारा calculate
            for p in set(affected_products):
                recalculate_virtual_stock(p)

            # ✅ Order status update
            order.status = "HOLD"
            order.notes = request.data.get("notes", order.notes)
            order.save()

        return Response(
            {"message": "Order put on HOLD and stock restored."},
            status=200
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=400)


@api_view(['POST'])
def reject_order(request, order_id):
    try:
        crm_user = request.user
        order = get_object_or_404(SSOrder, id=order_id, assigned_crm=crm_user)

        # वो snapshots लो जो पहले order verify/forward होने पर बने थे
        pending_snapshots = PendingOrderItemSnapshot.objects.filter(order=order)
        affected_products = [snap.product for snap in pending_snapshots]

        with transaction.atomic():

            # ✅ पहले snapshots delete — ये stock restore का trigger है
            pending_snapshots.delete()

            # ✅ हर product का virtual stock दोबारा calculate
            for p in set(affected_products):
                recalculate_virtual_stock(p)

            # ✅ Order status update
            order.status = "REJECTED"
            order.notes = request.data.get("notes", order.notes)
            order.save()

        return Response(
            {"message": "Order  Reject and stock restored."},
            status=200
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=400)


class CRMOrderListView(ListAPIView):
    serializer_class = SS_to_CRM_Orders
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        status_filter = self.request.query_params.get("status", "PENDING")

        base_queryset = (
            SSOrder.objects
            .filter(status=status_filter)
            .exclude(crm_verified_versions__isnull=False)
            .select_related(
                "ss_user",
                "assigned_crm"
            )
            .prefetch_related(
                "items__product"
            )
            .order_by("-created_at")
        )

        # 🔹 Admin → all orders
        if user.is_staff or user.is_superuser:
            return base_queryset[:20]

        # 🔹 CRM → only assigned orders
        return base_queryset.filter(assigned_crm=user)[:25]


class CRMOrderVerifyView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, order_id):
        try:
            crm_user = request.user
            data = request.data

            original_order = get_object_or_404(
                SSOrder, id=order_id, assigned_crm=crm_user
            )

            # Prevent duplicate verification
            if CRMVerifiedOrder.objects.filter(original_order=original_order).exists():
                return Response(
                    {"error": "Order already verified"}, status=status.HTTP_400_BAD_REQUEST
                )

            with transaction.atomic():
                # Create CRM verification record
                crm_order = CRMVerifiedOrder.objects.create(
                    original_order=original_order,
                    crm_user=crm_user,
                    status=data["status"],
                    notes=data.get("notes", ""),
                    total_amount=data.get("total_amount", 0),
                    dispatch_location=data.get("dispatch_location"),
                )

                # Original SS items map
                ss_items = SSOrderItem.objects.filter(order=original_order).select_related("product")
                ss_map = {
                    i.product.product_id: {
                        "product_obj": i.product,
                        "quantity": i.quantity,
                        "price": i.price,
                        "ss_virtual_stock": i.ss_virtual_stock,
                    }
                    for i in ss_items
                }

                # If whole order rejected -> mark items rejected
                if data["status"] == "REJECTED":
                    for pid, info in ss_map.items():
                        CRMVerifiedOrderItem.objects.create(
                            crm_order=crm_order,
                            product=info["product_obj"],
                            quantity=info["quantity"],
                            price=Decimal(info.get("price") or 0),
                            ss_virtual_stock=info.get("ss_virtual_stock", 0),
                            is_rejected=True,
                        )

                    # delete snapshots (no reservation anymore)
                    pending_snapshots = PendingOrderItemSnapshot.objects.filter(order=original_order)
                    affected_products = [snap.product for snap in pending_snapshots]
                    pending_snapshots.delete()

                    for p in set(affected_products):
                        recalculate_virtual_stock(p)

                else:
                    # Partial / Full approval
                    payload_items = data.get("items", [])
                    kept_products = set()
                    approved_map = {}
                    affected_products = set()

                    for item in payload_items:
                        product = Product.objects.get(product_id=item["product"])
                        try:
                            qty = int(item.get("quantity", 0))
                        except (TypeError, ValueError):
                            qty = 0

                        kept_products.add(product.product_id)
                        approved_map[product.product_id] = qty

                        ss_item = SSOrderItem.objects.filter(order=original_order, product=product).first()

                        # Safe price conversion
                        raw_price = item.get("price", 0)
                        try:
                            price_value = Decimal(raw_price) if raw_price not in [None, "", "null"] else Decimal(0)
                        except (InvalidOperation, TypeError, ValueError):
                            price_value = Decimal(0)

                        # ✅ अगर SSOrderItem नहीं मिला, तो product.virtual_stock का इस्तेमाल करो
                        if ss_item:
                            ss_virtual_stock_value = ss_item.ss_virtual_stock
                        else:
                            ss_virtual_stock_value = product.virtual_stock or 0

                        CRMVerifiedOrderItem.objects.create(
                            crm_order=crm_order,
                            product=product,
                            quantity=qty,
                            price=price_value,
                            ss_virtual_stock=ss_virtual_stock_value,
                            is_rejected=False,
                        )


                    # Products removed by CRM are considered rejected
                    deleted_products = set(ss_map.keys()) - kept_products
                    for pid in deleted_products:
                        info = ss_map[pid]
                        price_value = Decimal(info.get("price") or 0)
                        CRMVerifiedOrderItem.objects.create(
                            crm_order=crm_order,
                            product=info["product_obj"],
                            quantity=info["quantity"],
                            price=price_value,
                            ss_virtual_stock=info.get("ss_virtual_stock", 0),
                            is_rejected=True,
                        )

                    # Update PendingOrderItemSnapshot
                    for pid, approved_qty in approved_map.items():
                        prod_obj = ss_map.get(pid, {}).get("product_obj")
                        if not prod_obj:
                            prod_obj = Product.objects.get(product_id=pid)
                        PendingOrderItemSnapshot.objects.update_or_create(
                            order=original_order,
                            product=prod_obj,
                            defaults={"quantity": approved_qty},
                        )
                        affected_products.add(prod_obj)

                    for pid in deleted_products:
                        info = ss_map[pid]
                        PendingOrderItemSnapshot.objects.filter(order=original_order, product=info["product_obj"]).delete()
                        affected_products.add(info["product_obj"])

                    # Recalculate virtual stock for all affected products
                    for p in set(affected_products):
                        recalculate_virtual_stock(p)

                # Update order status
                original_order.status = data["status"]
                original_order.save(update_fields=["status"])

            return Response(
                {
                    "message": "Order verified successfully",
                    "crm_order": CRMVerifiedOrderSerializer(crm_order).data,
                },
                status=status.HTTP_201_CREATED,
            )

        except Product.DoesNotExist:
            return Response({"error": "Product not found"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class CRMOrderBulkDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        order_ids = request.data.get("order_ids", [])

        # ✅ ONLY ADMIN CAN DELETE
        if user.role != "ADMIN":
            return Response(
                {"error": "You are not allowed to delete orders."},
                status=status.HTTP_403_FORBIDDEN
            )

        if not order_ids:
            return Response(
                {"error": "No orders selected"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ✅ ADMIN can delete ANY order
        orders = SSOrder.objects.filter(id__in=order_ids)

        if not orders.exists():
            return Response(
                {"error": "No valid orders found"},
                status=status.HTTP_404_NOT_FOUND
            )

        affected_products = []

        with transaction.atomic():
            for order in orders:
                snapshots = PendingOrderItemSnapshot.objects.filter(order=order)
                affected_products.extend([s.product for s in snapshots])
                snapshots.delete()
                order.delete()

            # ✅ Recalculate stock once per product
            for product in set(affected_products):
                recalculate_virtual_stock(product)

        return Response(
            {"message": f"{orders.count()} orders permanently deleted"},
            status=status.HTTP_200_OK
        )



@api_view(["GET"])
def list_orders_by_role(request):
    user = request.user
    order_id = request.GET.get("order_id")
    party_name = request.GET.get("party_name")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    # 🟦 Base Query
    if user.role == "ADMIN":
        orders = SSOrder.objects.all()
    elif user.role == "CRM":
        orders = SSOrder.objects.filter(assigned_crm=user)
    elif user.role == "SS":
        orders = SSOrder.objects.filter(ss_user=user)
    else:
        orders = SSOrder.objects.none()

    # 🟦 Filters
    if order_id:
        orders = orders.filter(order_id__icontains=order_id)

    if party_name:
        orders = orders.filter(ss_user__party_name__icontains=party_name)

    if from_date:
        orders = orders.filter(created_at__date__gte=from_date)

    if to_date:
        orders = orders.filter(created_at__date__lte=to_date)

    # 🟦 Default limit (latest 50)
    if not (from_date or to_date or order_id or party_name):
        orders = orders.order_by("-created_at")[:15]
    else:
        orders = orders.order_by("-created_at")

    serializer = SSOrderSerializerTrack(orders, many=True)
    return Response(serializer.data)


class CombinedOrderTrackView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, order_id):
        user = request.user

        try:
            order = SSOrder.objects.get(order_id=order_id)
        except SSOrder.DoesNotExist:
            return Response({"error": "Order not found"}, status=404)

        # ✅ CRM apne assigned orders hi dekhega
        if not (user.is_staff or user.is_superuser):
            if order.assigned_crm != user and order.ss_user != user:
                return Response({"error": "Not authorized"}, status=403)


        data = CombinedOrderTrackSerializer(order).data
        return Response(data, status=200)


class CRMVerifiedOrderHistoryView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = CRMVerifiedOrderListSerializer

    def get_queryset(self):
        user = self.request.user

        qs = CRMVerifiedOrder.objects.select_related(
            "original_order",
            "original_order__ss_user",
            "crm_user"
        )

        if not user.is_staff and not user.is_superuser:
            qs = qs.filter(crm_user=user)

        q = self.request.query_params.get("q")
        punched_param = self.request.query_params.get("punched")

        if q:
            qs = qs.filter(
                Q(id__icontains=q) |                          # 9898
                Q(original_order__order_id__icontains=q)      # ORD-6CF49D38
            )

        if punched_param is not None:
            qs = qs.filter(punched=(punched_param.lower() == "true"))
        else:
            qs = qs.filter(punched=False)

        return qs.order_by("-verified_at")[:10]


class UpdateOrderStatusView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        crm_order = get_object_or_404(CRMVerifiedOrder, pk=pk)
        new_status = request.data.get("status")
        notes = request.data.get("notes")

        if new_status not in ["HOLD", "APPROVED", "REJECTED"]:
            return Response({"detail": "Invalid status"}, status=drf_status.HTTP_400_BAD_REQUEST)

        crm_order.status = new_status
        crm_order.notes = notes if new_status in ["HOLD", "REJECTED"] else None
        crm_order.save(update_fields=["status", "notes"])

        ss_order = crm_order.original_order
        ss_order.status = new_status
        ss_order.notes = notes if new_status in ["HOLD", "REJECTED"] else None
        ss_order.save(update_fields=["status", "notes"])

        return Response({
            "detail": "Status updated successfully",
            "status": new_status,
            "notes": crm_order.notes
        })

@api_view(['POST'])
def punch_order_to_sheet(request):
    try:
        data = request.data

        order_id = data.get("order_id")
        ss_party_name = data.get("ss_party_name")
        crm_name = data.get("crm_name")
        ss_id = data.get("id")
        dispatch_location = data.get("dispatch_location", "")
        items = data.get("items", [])
        is_single_row = data.get("is_single_row", False)

        if not order_id or not items:
            return Response({"error": "Missing order_id or items"}, status=400)

        # 🔐 Bulk punch only → prevent double punch
        order_obj = None
        if not is_single_row:
            order_obj = CRMVerifiedOrder.objects.filter(
                original_order__order_id=order_id,
                punched=False
            ).first()

            if not order_obj:
                return Response(
                    {"error": "Order already punched"},
                    status=400
                )

        ist_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        rows = [
            [
                item.get("product_name", ""),
                int(item.get("quantity", 0)),
                ss_party_name,
                ss_id,
                crm_name,
                item.get("id", ""),
                ist_timestamp,
                order_id,
                dispatch_location,
            ]
            for item in items
        ]

        # ✅ Always write to sheet
        write_to_sheet(
            settings.SHEET_ID_NEW,
            "order_data_from_app",
            rows
        )

        # ✅ Mark punched ONLY for bulk punch
        if not is_single_row and order_obj:
            order_obj.punched = True
            order_obj.dispatch_location = dispatch_location
            order_obj.save(update_fields=["punched", "dispatch_location"])

        return Response({
            "success": True,
            "message": f"{len(rows)} rows punched successfully",
            "single_row": is_single_row
        })

    except Exception as e:
        logger.error("🚨 Error in punch_order_to_sheet", exc_info=True)
        return Response(
            {"success": False, "error": str(e)},
            status=500
        )


class AddItemToCRMVerifiedOrderView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        crm_order = get_object_or_404(CRMVerifiedOrder, pk=pk)

        product_id = request.data.get("product_id")
        raw_qty = request.data.get("quantity")
        raw_price = request.data.get("price", 0)

        if not product_id or not raw_qty:
            return Response(
                {"error": "Product ID and quantity are required."},
                status=400
            )

        # ✅ SAFE quantity
        try:
            quantity = int(raw_qty)
        except (TypeError, ValueError):
            return Response({"error": "Invalid quantity."}, status=400)

        # ✅ SAFE price (🔥 MAIN FIX)
        try:
            price = Decimal(raw_price)
        except (InvalidOperation, TypeError):
            price = Decimal("0")

        product = get_object_or_404(Product, product_id=product_id)

        # ✅ Prevent duplicate product
        if CRMVerifiedOrderItem.objects.filter(
            crm_order=crm_order, product=product
        ).exists():
            return Response(
                {"error": "This product is already added in this order."},
                status=400
            )

        # ✅ Stock logic
        ss_stock = getattr(product, "ss_virtual_stock", 0)
        virtual_stock = getattr(product, "virtual_stock", 0)

        new_item = CRMVerifiedOrderItem.objects.create(
            crm_order=crm_order,
            product=product,
            quantity=quantity,
            price=price,   # 🔥 ALWAYS Decimal 0 or valid
            ss_virtual_stock=ss_stock if ss_stock > 0 else virtual_stock
        )

        # ✅ Total calculation safe
        crm_order.total_amount += price * Decimal(quantity)
        crm_order.save(update_fields=["total_amount"])

        return Response(
            {
                "message": "Product added successfully!",
                "item_id": new_item.id,
                "product_name": product.product_name,
                "quantity": quantity,
                "price": price,
            },
            status=201
        )


class CRMVerifiedItemUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        """
        Update quantity or price of a verified order item
        """
        try:
            item = CRMVerifiedOrderItem.objects.get(pk=pk)
        except CRMVerifiedOrderItem.DoesNotExist:
            return Response({"error": "Item not found"}, status=status.HTTP_404_NOT_FOUND)

        quantity = request.data.get("quantity")
        price = request.data.get("price")

        if quantity is not None:
            item.quantity = quantity
        if price is not None:
            item.price = price

        item.save()
        return Response({"message": "Item updated successfully"})


class CRMVerifiedItemDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            item = CRMVerifiedOrderItem.objects.get(pk=pk)
            item.delete()
            return Response({"message": "Item deleted successfully"}, status=status.HTTP_200_OK)
        except CRMVerifiedOrderItem.DoesNotExist:
            return Response({"error": "Item not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(["POST"])
def submit_meet_form(request):
    try:
        data = request.data

        business_name = data.get("business_name", "")
        person_name = data.get("person_name", "")
        phone = data.get("phone", "")
        district = data.get("district", "")

        # validation
        if not person_name or not phone:
            return Response({"error": "Person name and phone required"}, status=400)

        # IST timestamp
        ist_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # row for google sheet
        row = [
            business_name,
            person_name,
            phone,
            district,
            ist_timestamp
        ]

        # sheet name → abc
        write_to_sheet(
            settings.SHEET_ID_NEW,
            "abc",
            [row]
        )

        return Response({"success": True, "message": "Form submitted successfully"})

    except Exception as e:
        logger.error(f"Meet form error: {str(e)}", exc_info=True)
        return Response({"error": "Internal Server Error"}, status=500)


@api_view(["POST"])
def submit_dealer_list(request):
    try:
        dealers = request.data.get("dealers", [])

        if not dealers:
            return Response({"error": "No dealers found"}, status=400)

        rows = []

        for d in dealers:
            rows.append([
                d.get("dealer_name", ""),
                d.get("shop_name", ""),
                d.get("mobile", ""),
                d.get("block", ""),
                d.get("district", ""),
                d.get("quantity", ""),
                d.get("pin_code", ""),

                # ⭐ NEW SPECIAL FIELDS ⭐
                d.get("designation", ""), 
                d.get("your_name", ""),
                d.get("super_stockist_name", ""),
                d.get("super_stockist_crm", ""),
                d.get("distributor_name", ""),

                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # timestamp
            ])

        # Google Sheet Page Name
        write_to_sheet(
            settings.SHEET_ID_NEW,
            "abc",
            rows
        )

        return Response({"success": True, "message": "Dealers submitted successfully"})

    except Exception as e:
        logger.error(f"Dealer submit error: {str(e)}", exc_info=True)
        return Response({"error": "Internal Server Error"}, status=500)



class DispatchOrderListView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = DispatchOrderSerializer

    def get_queryset(self):
        qs = DispatchOrder.objects.all()

        from_date = self.request.query_params.get("from")
        to_date = self.request.query_params.get("to")

        if from_date:
            qs = qs.filter(order_packed_time__date__gte=from_date)

        if to_date:
            qs = qs.filter(order_packed_time__date__lte=to_date)

        qs = qs.order_by("-order_packed_time")

        # ✅ ONLY limit when NO filters
        if not from_date and not to_date:
            return qs[:10]

        return qs


class DeleteAllDispatchOrders(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        count, _ = DispatchOrder.objects.all().delete()
        return Response(
            {
                "message": "सभी dispatch orders delete हो गए",
                "deleted_count": count
            },
            status=status.HTTP_200_OK
        )
    
class DeleteSelectedDispatchOrders(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ids = request.data.get("ids", [])
        count, _ = DispatchOrder.objects.filter(id__in=ids).delete()
        return Response(
            {"deleted_count": count},
            status=200
        )


class DownloadDispatchExcel(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dispatch Orders"

        # Header
        ws.append([
            "order_id",
            "product",
            "quantity",
            "order_packed_time",  # optional
        ])

        # Example row (optional)
        ws.append([
            "ORD-12345",
            "Product A",
            10,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=dispatch_orders.xlsx"
        wb.save(response)
        return response


class UploadDispatchExcel(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")

        if not file:
            return Response(
                {"error": "Excel file required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        created = 0

        for row in ws.iter_rows(min_row=2, values_only=True):

            # ✅ SAFE INDEX READ (NO UNPACKING)
            order_id = row[0] if len(row) > 0 else None
            product = row[1] if len(row) > 1 else None
            quantity = row[2] if len(row) > 2 else None
            packed_time = row[3] if len(row) > 3 else None

            if not order_id or not product or not quantity:
                continue

            # 🔥 DATE HANDLE
            if packed_time:
                if isinstance(packed_time, datetime):
                    final_time = packed_time
                elif isinstance(packed_time, str):
                    try:
                        final_time = datetime.strptime(
                            packed_time.strip(),
                            "%d-%m-%Y %H:%M"
                        )
                    except ValueError:
                        final_time = timezone.now()
                else:
                    final_time = timezone.now()
            else:
                final_time = timezone.now()

            DispatchOrder.objects.create(
                order_id=str(order_id).strip(),
                product=str(product).strip(),
                quantity=int(quantity),
                order_packed_time=final_time
            )

            created += 1

        return Response(
            {"message": "Upload successful", "created": created},
            status=status.HTTP_201_CREATED
        )
