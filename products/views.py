import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from rest_framework.pagination import PageNumberPagination
from rest_framework import status
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework.decorators import api_view
from django.db import transaction
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from .models import  Product, SaleName, Scheme
from .serializers import (  ProductSerializer, SaleNameSerializer,SchemeSerializer, ProductWithSaleNameSerializer)


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all().order_by('product_id')
    serializer_class = ProductSerializer
    lookup_field = 'product_id'

    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser])
    def upload_image(self, request, product_id=None):
        product = self.get_object()
        image_file = request.FILES.get('image')
        if not image_file:
            return Response({"error": "No image provided"}, status=400)
        product.image = image_file
        product.save()
        return Response({'status': 'image uploaded', 'url': product.image.url})

    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser])
    def upload_image2(self, request, product_id=None):
        product = self.get_object()
        image_file = request.FILES.get('image2')
        if not image_file:
            return Response({"error": "No image2 provided"}, status=400)
        product.image2 = image_file
        product.save()
        return Response({'status': 'image2 uploaded', 'url': product.image2.url})


class ProductBulkTemplateDownload(APIView):
    def get(self, request):
        # Template columns
        columns = ["product_id", "product_name", "sub_category", "cartoon_size", "guarantee", "price", "moq", "rack_no"]
        df = pd.DataFrame(columns=columns)

        # Excel response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="product_template.xlsx"'
        df.to_excel(response, index=False)
        return response


class ProductBulkUpload(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)

            required_columns = {"product_id", "product_name", "sub_category", "cartoon_size", "guarantee", "price", "moq", "rack_no"}
            if not required_columns.issubset(df.columns):
                return Response({"error": "Invalid file format"}, status=status.HTTP_400_BAD_REQUEST)

            created_count = 0
            updated_count = 0

            for _, row in df.iterrows():
                product_id = row["product_id"]

                # Clean cartoon_size
                cartoon_size_raw = row.get("cartoon_size", "")
                if pd.notnull(cartoon_size_raw):
                    if isinstance(cartoon_size_raw, float) and cartoon_size_raw.is_integer():
                        cartoon_size_str = str(int(cartoon_size_raw))  # e.g., 200.0 → "200"
                    else:
                        cartoon_size_str = str(cartoon_size_raw)       # e.g., 250.5 or "text"
                else:
                    cartoon_size_str = ""

                # Prepare defaults
                defaults = {
                    "product_name": row.get("product_name", ""),
                    "sub_category": row.get("sub_category", ""),
                    "cartoon_size": cartoon_size_str,
                    "guarantee": row.get("guarantee", ""),
                    "price": str(row.get("price", "")) if pd.notnull(row.get("price")) else "",
                    "moq": int(row.get("moq", 0)) if pd.notnull(row.get("moq")) else 0,
                    "rack_no": row.get("rack_no", ""),
                }

                obj, created = Product.objects.update_or_create(
                    product_id=product_id,
                    defaults=defaults
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            return Response({
                "message": "Bulk upload completed",
                "created": created_count,
                "updated": updated_count
            })

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class SaleNamePagination(PageNumberPagination):
    page_size = 10  # default
    page_size_query_param = 'page_size'


class SaleNameViewSet(viewsets.ModelViewSet):
    queryset = SaleName.objects.select_related('product').all().order_by('product_id')
    serializer_class = SaleNameSerializer
    pagination_class = SaleNamePagination

    def get_queryset(self):
        queryset = super().get_queryset()
        product_id = self.request.query_params.get('product_id')  # ← छोटा bug था '-product_id'
        if product_id:
            queryset = queryset.filter(product__product_id=product_id)
        return queryset

    @action(detail=False, methods=['delete'], url_path='delete-by-product/(?P<product_id>[^/.]+)')
    def delete_by_product(self, request, product_id=None):
        count, _ = SaleName.objects.filter(product__product_id=product_id).delete()
        return Response({"message": f"{count} sale names deleted successfully."}, status=status.HTTP_200_OK)

    # 🆕 Delete All SaleNames
    @action(detail=False, methods=['delete'], url_path='delete-all')
    def delete_all(self, request):
        count, _ = SaleName.objects.all().delete()
        return Response({"message": f"{count} sale names deleted successfully."}, status=status.HTTP_200_OK)


class SaleNameBulkUploadView(APIView):
    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file uploaded."}, status=400)

        try:
            df = pd.read_excel(file) if file.name.endswith(".xlsx") else pd.read_csv(file)
            if not {"product_id", "sale_name"}.issubset(df.columns):
                return Response({"error": "File must contain 'product_id' and 'sale_name' columns"}, status=400)

            skipped_ids = []
            saved_count = 0

            with transaction.atomic():
                for _, row in df.iterrows():
                    product_id = row.get("product_id")
                    sale_name = row.get("sale_name")

                    if pd.isna(product_id) or str(product_id).strip() == "":
                        skipped_ids.append("Missing")
                        continue

                    product = Product.objects.filter(product_id=product_id).first()
                    if product:
                        SaleName.objects.create(product=product, sale_name=sale_name)
                        saved_count += 1
                    else:
                        skipped_ids.append(str(product_id))

            return Response({
                "message": f"{saved_count} sale names uploaded successfully.",
                "skipped_ids": skipped_ids,
                "skipped_count": len(skipped_ids)
            }, status=201)

        except Exception as e:
            return Response({"error": str(e)}, status=500)


class SchemeViewSet(viewsets.ModelViewSet):
    queryset = Scheme.objects.all().order_by('-id')
    serializer_class = SchemeSerializer
    permission_classes = [IsAuthenticated]


@api_view(['GET'])
def get_all_products_with_salenames(request):
    products = Product.objects.filter(is_active=True).prefetch_related('sale_names').order_by('product_id')
    serializer = ProductWithSaleNameSerializer(products, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def get_virtual_stock(request):
    products = Product.objects.filter(is_active=True).values("product_id", "virtual_stock")
    return Response(products)


@api_view(['GET'])
def get_inactive_products(request):
    products = Product.objects.filter(is_active=False).prefetch_related('sale_names').order_by('product_id')
    serializer = ProductWithSaleNameSerializer(products, many=True)
    return Response(serializer.data)



@api_view(["GET"])
def export_products_excel(request):
    # ✅ Workbook create
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # ✅ Columns (same order you requested)
    columns = [
        "Product_id",
        "Sub_category",
        "Product_name",
        "Cartoon",
        "Guarantee",
        "Price",
        "MOQ",
        "Rack"
    ]

    ws.append(columns)

    # ✅ Add product rows
    for p in Product.objects.all().order_by("product_id"):
        ws.append([
            p.product_id,
            p.sub_category or "",
            p.product_name,
            p.cartoon_size or "",
            p.guarantee or "",
            p.price or "",
            p.moq or "",
            p.rack_no or "",
        ])

    # ✅ Set column width auto-adjust (optional but professional)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # ✅ Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="products.xlsx"'

    wb.save(response)
    return response
